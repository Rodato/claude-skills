#!/usr/bin/env python3
"""Helpers mínimos reutilizables para los informes Excel de AMA.

Extraídos/inspirados del código real (make_excel.py, informe_cobertura_excel.py,
report_bot.py, user_report.py). No es un framework: copiá lo que necesites y ajustá los
colores del header a la familia del informe que estés replicando (ver
references/estilo-openpyxl.md). Comentarios en español, código en inglés. python3.

Uso típico:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    write_header_row(ws, 1, ["Fecha", "Escuela", "Diferencia"])
    # ... escribir filas, aplicar semáforo por celda ...
    autosize_by_content(ws)
    wb.save("data/reports/mi_informe_2026-07-10.xlsx")
"""
from __future__ import annotations

import io

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Semáforo (hex constantes entre informes) ──────────────────────────────────
GREEN = "C6EFCE"        # bien / diferencia positiva / % cobertura ≥ 70
YELLOW = "FFEB9C"       # atención / diferencia negativa (faltantes) / 40 ≤ % < 70
RED = "FFC7CE"          # crítico / % cobertura < 40
NAME_MATCH = "FCE4D6"   # naranja: fila con match por nombre → posible ID incorrecto
REVISAR = "FFF2CC"      # naranja pálido: fila de baja confianza a verificar a mano

# ── Header (teal Lago Agrio; cambialo por la paleta del informe que copiás) ────
HDR_FILL = "1F4E5F"
HDR_FONT_COLOR = "FFFFFF"


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def thin_border(color: str = "D0D0D0") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Semáforos ─────────────────────────────────────────────────────────────────

def diff_fill(value) -> str | None:
    """Semáforo a): color por signo de la Diferencia (= Kobo − Asistencia).

    Amarillo si negativo (asistencia > Kobo → faltan encuestas), verde si positivo
    (excedentes), None si es cero o no numérico (dejá el color base/zebra).
    """
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v < 0:
        return YELLOW
    if v > 0:
        return GREEN
    return None


def pct_fill(p) -> str | None:
    """Semáforo b): color por umbral de % (cobertura/avance). 70 / 40."""
    if p is None:
        return None
    if p >= 70:
        return GREEN
    if p >= 40:
        return YELLOW
    return RED


# ── Escritura de celdas ───────────────────────────────────────────────────────

def write_header_row(ws, row: int, cols, fill_hex: str = HDR_FILL) -> None:
    """Escribe una fila de encabezado estilizada (blanco sobre color, centrado, borde)."""
    border = thin_border()
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill(fill_hex)
        cell.font = Font(bold=True, color=HDR_FONT_COLOR, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 22


def write_text_cell(ws, row: int, col: int, value) -> None:
    """Escribe un valor forzándolo a TEXTO (teléfono/documento: no perder dígitos)."""
    cell = ws.cell(row=row, column=col, value="" if value is None else str(value))
    cell.number_format = "@"


def autosize_by_content(ws, max_width: int = 50, pad: int = 2) -> None:
    """Ajusta el ancho de cada columna al contenido más largo, con tope."""
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max_len + pad, max_width)


def set_col_widths(ws, widths) -> None:
    """Fija anchos explícitos (lista de números, una por columna desde A)."""
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ── Hoja-leyenda "Cómo leer este reporte" ─────────────────────────────────────

def write_legend_sheet(wb, title: str, sections, sheet_name: str = "Cómo leer este reporte"):
    """Crea la hoja-leyenda (primera del libro).

    sections: lista de (encabezado, [líneas]). Cada línea es una viñeta bajo su encabezado.
    Apaga la grilla y usa una sola columna ancha, como report_bot.py.
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14, color="1F4E79")
    ws.append([])

    for header, lines in sections:
        ws.append([header])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, color="2E75B6")
        for line in lines:
            ws.append(["  • " + line])
            ws.cell(row=ws.max_row, column=1).font = Font(size=10)
        ws.append([])

    ws.column_dimensions["A"].width = 90
    return ws


# ── Bytes en memoria para Streamlit ───────────────────────────────────────────

def workbook_to_bytes(wb) -> bytes:
    """Serializa el Workbook a bytes para st.download_button (sin tocar disco)."""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
